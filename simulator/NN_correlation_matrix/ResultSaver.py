import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl import Workbook, load_workbook
import torch

#saving_path = "C:\\Users\\Ido\\PycharmProjects\\metabolic_pathways\\simulator\\NN_correlation_matrix\\Results\\"
saving_path = "Results/"
excel_path = saving_path + "result.xlsx"

class ResultSaver:
    def __init__(self, minibatch_size, step, dataset_size, epochs, reactions_count):
        self.minibatch_size = minibatch_size
        self.step = step
        self.dataset_size = dataset_size
        self.epochs = epochs
        self.reactions_count = reactions_count
        wb = Workbook()
        wb.save(excel_path)

    def write_date_final_epoch(self, gan_structure, model_output, real_output, loss_list):
        wb = load_workbook(excel_path)
        sheet_name = gan_structure.get_name()
        filepath = saving_path + 'final_images\\' + sheet_name + '_final_image.png'
        color = "YlGnBu"

        # Create the matrix for yc result
        plt.close()
        plt.title('correlation matrix for model ' + gan_structure.get_name())
        npimg = model_output.detach().cpu().numpy()
        plt.imshow(npimg, cmap=color)
        plt.savefig(filepath, dpi=50)

        # Create the original matrix graph
        plt.close()
        plt.title('original matrix')
        npimg = real_output.cpu().numpy()
        plt.imshow(npimg, cmap=color) #Greys
        plt.savefig(saving_path + 'original_matrix_images\\' + sheet_name + '.png', dpi=50)

        # plot the graph of epchos-MSE
        plt.close()
        plt.title("Epochs with MSE")
        plt.plot(range(self.epochs), loss_list)
        plt.xlabel('Epochs')
        plt.ylabel('MSE loss')
        plt.savefig(saving_path + 'graphs\\' + sheet_name + '.png', dpi=50)

        # Add photos to excel sheet
        # Activate worksheet
        active = wb[sheet_name]

        # Insert plot into worksheet
        # Select active sheet and cell reference
        img = Image(filepath)
        active.add_image(img, 'I1')

        img_orginial_matrix = Image(saving_path + 'original_matrix_images\\' + sheet_name + '.png')
        active.add_image(img_orginial_matrix, 'N1')

        img_grraph = Image(saving_path + 'graphs\\' + sheet_name + '.png')
        active.add_image(img_grraph, 'I15')

        # Save workbook
        wb.save(excel_path)

    def write_data_epoch(self, gan_structure, loss, ploss, rloss, epoch, time):
        wb = load_workbook(excel_path)
        sheet_name = gan_structure.get_name()
        if epoch == 0:
            wb.create_sheet(sheet_name, 0)
            active = wb[sheet_name]
            line_append = ["epoch", "time", "mse-loss", "ploss", "rloss", "minibatch_size", "learning rate",
                           "dataset size"]
            active.append(line_append)

        active = wb[sheet_name]
        line_append = [epoch, time, loss.item(), ploss, rloss, self.minibatch_size, self.step, self.dataset_size]
        active.append(line_append)

        # Save workbook to write
        wb.save(excel_path)

    def write_data_summary(self, structure_list):
        wb = load_workbook(excel_path)
        sheet_name = "Summary"
        wb.create_sheet(sheet_name, 0)
        active = wb[sheet_name]
        line_append = ["model", "mse loss"]  # "input error", "output error"]
        for i in range(self.reactions_count):
            line_append.append("Reaction " + str(i + 1) + " input error")
            line_append.append("Reaction " + str(i + 1) + " output error")
        active.append(line_append)
        for structure in structure_list:
            in_err_dict, out_err_dict = structure.get_score()
            line_append = [structure.get_name(), structure.get_last_mse_loss()]
                           #structure.get_last_mse_loss().detach().item()]  # ,in_err + '%', out_err + '%']
            for i in range(self.reactions_count):
                in_err = in_err_dict[i]
                out_err = out_err_dict[i]
                line_append.append(in_err + '%')
                line_append.append(out_err + '%')
            active.append(line_append)

        wb.save(excel_path)

    def save_reactions(self, gan_structure):
        wb = load_workbook(excel_path)
        sheet_name = gan_structure.get_name()
        active = wb[sheet_name]

        # add original and predicted
        original_lst, predicted_lst = gan_structure.get_reactions_lists()
        active.append([""])  # blank line
        active.append(["original reactions: ", str(original_lst)])
        active.append(["predicted reactions: ", str(predicted_lst)])

        wb.save(excel_path)

    def save_two_random(self, y1, y2, itreation, mse_loss):
        two_random_path = saving_path + "\\result_random.xlsx"
        sheet_name = str(itreation)
        if itreation == 0:
            wb = Workbook()
            wb.save(two_random_path)
        wb = load_workbook(two_random_path)
        wb.create_sheet(sheet_name, 0)

        # Create first matrix
        plt.close()
        plt.title('First Matrix')
        npimg = y1.cpu().numpy()
        plt.imshow(npimg, cmap="YlGnBu")  # YlGnBu
        plt.savefig(saving_path + 'original_matrix_images\\' + sheet_name + '1.png', dpi=50)

        # Create second matrix
        plt.close()
        plt.title('First Matrix')
        npimg = y2.cpu().numpy()
        plt.imshow(npimg, cmap="YlGnBu")  # YlGnBu
        plt.savefig(saving_path + 'original_matrix_images\\' + sheet_name + '2.png', dpi=50)

        # Add photos to excel sheet
        # Activate worksheet
        active = wb[sheet_name]
        active.append([mse_loss])

        # Insert plot into worksheet
        # Select active sheet and cell reference
        img = Image(saving_path + 'original_matrix_images\\' + sheet_name + '1.png')
        active.add_image(img, 'A3')

        img_second_matrix = Image(saving_path + 'original_matrix_images\\' + sheet_name + '2.png')
        active.add_image(img_second_matrix, 'F3')

        wb.save(two_random_path)

    def save_model(self, model, gan_structure):
        name = gan_structure.get_name()
        #torch.save(model.state_dict(), saving_path + "\\Saved Models\\" + name + ".pt")
        torch.save(model, saving_path + "\\Saved Models\\" + name + ".pt")
