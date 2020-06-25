#Class to process results from the output of the trained model
#also in charge of providing the classifier with proper files
import os
import torch
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl import Workbook, load_workbook
import random
import xlrd
import csv


global device
device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")



class ResultProcessor:

    def __init__(self, m_count, minibatch_size, sub_min, sub_max):
        self.saving_path ="Results/"
        self.outputs_path = self.saving_path + "\\Saved Models\\Outputs"
        self.model = None
        self.m_count = m_count
        self.minibatch_size = minibatch_size
        self.sub_min = sub_min
        self.sub_max = sub_max
        self.wb = Workbook()


    def load_model(self, name):
        #model = Process_new(self.reactions_count, self.metabolites, self.m_count, self.m_count, self.iterations)
        #model.load_state_dict()
        model = torch.load(self.saving_path + "\\Saved Models\\" + name)
        model.eval()
        self.model = model


    def run_model(self, matrix_count):
        for i in range(matrix_count):
            reactions = None
            reaction_num = 20
            self.wb = Workbook()
            while reactions is None or len(reactions) < reaction_num:
                rand_x = torch.Tensor(size=(self.minibatch_size, self.m_count))
                rand_x.uniform_(self.sub_min, self.sub_max)
                rand_x = rand_x.to(device)

                # ym, yc = model(x)
                ym, yc = self.model(rand_x)
                self.draw_matrix(yc, i)

                if reactions is None:
                    reactions = self.model.get_reactions_tensor()
                else:
                    reactions = torch.cat([reactions, self.model.get_reactions_tensor()])
            self.write_reactions_data(reactions, i)
            self.create_semi_random_negative_instances(i)

            matrix = yc.detach().numpy()
            self.write_matrix_data(matrix, i)

    def run_model_for_fake(self, matrix_count, reactions, yc):
        for i in range(matrix_count):
            self.wb = Workbook()
            self.draw_matrix(yc, i)

            self.write_reactions_data(reactions, i)
            self.create_semi_random_negative_instances(i)

            matrix = yc.detach().numpy()
            self.write_matrix_data(matrix, i)

    #write matrix data to excel
    def write_matrix_data(self, matrix, index):
        #wb = Workbook()
        wb = self.wb
        sheet_name = "matrix"
        wb.create_sheet(sheet_name, 0)
        active = wb[sheet_name]

        #adding first raw of metabolites name
        metabolites = list(range(self.m_count))
        metabolites = ["X"] + metabolites #we insert 'X' so the table will look good
        line_append = metabolites
        active.append(line_append)

        for i in range(len(matrix)):
            line_append = [i] + matrix[i].tolist()
            active.append(line_append)

        wb.save(self.outputs_path + "\\Sample" + str(index) + "\\matrix.xlsx")

    def draw_matrix(self, model_output, matrix_id):
        semi_path = self.outputs_path + "\\Sample" + str(matrix_id)
        filepath = semi_path + "\\matrix_photo.png"
        if not os.path.exists(semi_path):
            os.makedirs(semi_path)
        color = "YlGnBu"
        dpi_size = 100
        wb = self.wb
        #wb = Workbook()
        sheet_name = "matrix_visual"

        # Create the matrix for yc result
        plt.close()
        plt.title('correlation matrix created by model')
        npimg = model_output.detach().cpu().numpy()
        plt.imshow(npimg, cmap=color)
        plt.savefig(filepath, dpi=dpi_size)

        # Add photos to excel sheet
        # Activate worksheet
        wb.create_sheet(sheet_name, 0)
        active = wb[sheet_name]

        # Insert plot into worksheet
        # Select active sheet and cell reference
        img = Image(filepath)
        active.add_image(img, 'A1')

        wb.save(self.outputs_path + "\\Sample" + str(matrix_id) + "\\result_p.xlsx")

    def get_saving_path(self):
        return self.saving_path

    def set_model(self, model):
        self.model = model

    #Write each reaction to a different excel file
    def write_reactions_data(self, reactions, matrix_index):
        reactions = reactions.round().int()
        print(reactions)
        for i in range(len(reactions)):
            reaction = reactions[i]
            wb = Workbook()
            sheet_name = "reaction" + str(i)
            wb.create_sheet(sheet_name, 0)
            active = wb[sheet_name]

            metabolic_list_num = []
            for j in range(self.m_count):
                if reaction[0][j] == 1:
                    metabolic_list_num.append(j)
                elif reaction[1][j] == 1: #doing else if because we don't want to add same metabolic twice
                    metabolic_list_num.append(j)

            metabolic_list = self.turn_metabolic_id_to_name(metabolic_list_num)

            line_append = metabolic_list_num #todo: change from metabolic_list_num to metabolic_list
            active.append(line_append)

            newpath = self.outputs_path + "\\Sample" + str(matrix_index) +"\\Positive\\"
            if not os.path.exists(newpath):
                os.makedirs(newpath)
            wb.save(newpath + "\\reaction" + str(i) + ".xlsx")

    #todo: complete this
    def turn_metabolic_id_to_name(self, metabolic_list_num):
        pass


    def create_semi_random_negative_instances(self, matrix_index):
        positive_reactions_list = self.read_all_positive_recations(matrix_index)

        count_negative_reactions = len(positive_reactions_list) #* 3
        negative_reactions_lst = []
        for i in range(count_negative_reactions):
            found_negative = False
            while(not found_negative):
                negative_rec = self.generate_random_reaction()
                if negative_rec not in positive_reactions_list: #and (negative_rec not in negative_reactions_lst): #does it matter if there are possible duplicates in the negative list?
                    negative_reactions_lst.append(negative_rec)
                    found_negative = True

                    metabolic_list_names = self.turn_metabolic_id_to_name(negative_rec)

                    line_append = negative_rec  # todo: change from metabolic_list_num to metabolic_list

                    wb = Workbook()
                    sheet_name = "neg_reaction" + str(i)
                    wb.create_sheet(sheet_name, 0)
                    active = wb[sheet_name]
                    active.append(line_append)
                    new_negative_path = self.outputs_path + "\\Sample" + str(matrix_index) + "\\Negative\\"
                    if not os.path.exists(new_negative_path):
                        os.makedirs(new_negative_path)
                    wb.save(new_negative_path + "\\neg_reaction" + str(i) + ".xlsx")



    #generate random reaction for the negative sample
    def generate_random_reaction(self):
        metabolic_list_num = []
        substrate = [random.randrange(0, 2, 1) for i in range(self.m_count)]
        product = [random.randrange(0, 2, 1) for i in range(self.m_count)]
        for j in range(self.m_count):
            if substrate[j] == 1:
                metabolic_list_num.append(j)
            elif product[j] == 1:  # doing else if because we don't want to add same metabolic twice
                metabolic_list_num.append(j)
        return metabolic_list_num

    def read_all_positive_recations(self, matrix_index):
        positive_directory = self.outputs_path + "\\Sample" + str(matrix_index) + "\\Positive\\"

        # read all positive reactions
        positive_reactions_list = []
        for filename in os.listdir(positive_directory):
            if filename.endswith(".xlsx"):
                wb = load_workbook(positive_directory + filename)
                sheet_name = filename.split(".")[0]  # remove the file extension
                worksheet = wb[sheet_name]
                first_row = worksheet[1]
                reaction = []
                for i in range(len(first_row)):
                    reaction.append(first_row[i].value)
                positive_reactions_list.append(reaction)

        return positive_reactions_list

    def convert_from_xlsx_to_cvs(self, matrices_num):
        for i in range(matrices_num):
            matrix_index = str(i)

            positive_directory = self.outputs_path + "\\Sample" + matrix_index + "\\Positive\\"
            positive_csv_path = self.outputs_path +  "\\Sample" + matrix_index + "\\Positive csv\\"
            if not os.path.exists(positive_csv_path):
                os.makedirs(positive_csv_path)

            for filename in os.listdir(positive_directory):
                wb = xlrd.open_workbook(positive_directory + os.sep + filename)
                sh = wb.sheet_by_name(filename.replace(".xlsx", ""))
                your_csv_file = open(positive_csv_path + filename.replace(".xlsx", ".csv"), 'w')
                wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)

                for rownum in range(sh.nrows):
                    wr.writerow(sh.row_values(rownum))

                your_csv_file.close()

            negative_directory = self.outputs_path + "\\Sample" + matrix_index + "\\Negative\\"
            negative_csv_path = self.outputs_path +  "\\Sample" + matrix_index + "\\Negative csv\\"
            if not os.path.exists(negative_csv_path):
                os.makedirs(negative_csv_path)

            for filename in os.listdir(negative_directory):
                wb = xlrd.open_workbook(negative_directory + os.sep + filename)
                sh = wb.sheet_by_name(filename.replace(".xlsx", ""))
                your_csv_file = open(negative_csv_path + filename.replace(".xlsx", ".csv"), 'w')
                wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)

                for rownum in range(sh.nrows):
                    wr.writerow(sh.row_values(rownum))

                your_csv_file.close()


